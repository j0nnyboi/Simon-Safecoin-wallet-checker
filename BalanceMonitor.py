
import requests
from time import gmtime, strftime,sleep
import datetime
from safecoin.keypair import Keypair
from safecoin.rpc.api import Client
from safecoin.rpc.types import MemcmpOpts
from safecoin.publickey import PublicKey

api_endpoint="https://api.mainnet-beta.safecoin.org"
client = Client(api_endpoint)

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter


def getBal(addr,client):
        try:
                addrBalance = int(client.get_balance(addr)['result']['value'])/1000000000
                print(addr," : ",addrBalance)
                return addrBalance
        except:
                return 0
        


Daypre = 99

while True:
        sleep(10)
        #day = strftime("%d", gmtime())#day check
        day = strftime("%H", gmtime())#hour check
        if(day != Daypre):#once an day we check if we need to update
            if(client.is_connected()):
                Daypre = day
                wb = load_workbook('Wallet_balances.xlsx')
                ws = wb.active
                
                for row1 in ws["1"]:#get first blank col for date and balances
                    ROW1 = row1
                colDate = "%s1" % (get_column_letter(ROW1.column + 1))
                           
                colA = ws['A']      
                for addr in colA:#get address in col A
                    #print(addr.value)
                    if(addr.value != "Wallet address"):
                        
                        addrBalance = getBal(addr.value,client)
                        oldcolval = "%s%s" % (get_column_letter(ROW1.column),addr.row)
                        #print("old value = ",ws[oldcolval].value)
                        if(ws[oldcolval].value != addrBalance):
                                ws.column_dimensions[get_column_letter(ROW1.column + 1)].width = 17
                                ws[colDate] = datetime.datetime.now()#add date to top
                                colval = "%s%s" % (get_column_letter(ROW1.column + 1),addr.row)
                                ws[colval] = addrBalance#add balance to row
                        else:
                                print("balance is the same as sheet")

                wb.save('Wallet_balances.xlsx')
                wb.close()
                        
            else:
                 client = Client(api_endpoint)

                
                    
                    

                
                    
                                        
                        
                        
                        

